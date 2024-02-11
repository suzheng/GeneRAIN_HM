from utils.utils import get_model, get_config, get_gene2idx
import os
import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt
from scipy.stats import mannwhitneyu, wilcoxon
import random
from collections import Counter
import copy
import torch
import os
import glob
import csv
import re
import numpy as np
# Set the directory where your CSV files are located
import pickle


def get_file_stat_dict_from_a_pickle_file(pickle_file):
    with open(pickle_file, 'rb') as handle:
        file_stat_dict = pickle.load(handle)
    return file_stat_dict
def extract_number_and_remaining(input_string):
    # Define a regular expression pattern to match 'e' followed by one or more digits at the end of the string
    pattern = r'e(\d+)$'

    # Use re.search to find the match
    match = re.search(pattern, input_string)

    # Check if a match is found
    if match:
        matched_number = match.group(1)  # Extract the number
        remaining_string = input_string[:match.start()-1]  # Extract the remaining string
        return int(matched_number), remaining_string
    else:
        return None, input_string  # No match found, return the original string

def get_gf_version_from_basename(basename):
    # pattern = r'__vs__(GF\d+L)'
    pattern = r'__vs__(G\w+)'
    match = re.search(pattern, basename)
    if match:
        return match.group(1)
    else:
        return None
        
def get_keys_from_file_name(file_name):
    exp_epoch, gene_set, basename = np.array(file_name.split("/"))[[-4, -2, -1]]
    epoch, exp = extract_number_and_remaining(exp_epoch)
    gf_version = get_gf_version_from_basename(basename)
    comp = None
    if "Null_vs_" in basename:
        pattern = r'(Null_vs_\w+)\.'
        # Use re.search to find the match
        match = re.search(pattern, basename)
        comp = match.group(1)
    else:
        comp = "Self_vs_" + gf_version
    if comp == "Null_vs_Self":
        comp = f"{comp}_{gf_version}"
    return (exp, epoch, gene_set, basename, gf_version, comp)





def get_naming_dict():
    config = get_config()
    naming_files = glob.glob(f"{config.proj_path}/data/meta/naming/naming.*.txt")
    naming_dict = {}
    for naming_file in naming_files:
        basename = os.path.basename(naming_file)
        key = basename.replace("naming.", "").replace(".txt", "")
        df = pd.read_csv(naming_file, sep='\t')
        naming_dict[key] = df
    return naming_dict
def get_genesets_of_interest_df(only_clustering_databases=True):
    # config = get_config()
    # genesets_of_interest_file = config.proj_path + "/data/external/Enrichr/092023/genesets_of_interest.tsv"
    # df = pd.read_csv(genesets_of_interest_file, sep='\t')
    
    # return df
    naming_dict = get_naming_dict()
    genesets_of_interest_df = naming_dict['Geneset']
    if only_clustering_databases:
        genesets_of_interest_df = genesets_of_interest_df[genesets_of_interest_df['Only_in_clf']==False]
    # genesets_of_interest_df['Gene_Set'] = genesets_of_interest_df['Database_ori_name']
    genesets_of_interest_df = genesets_of_interest_df.rename(columns={'Database_ori_name': 'Gene_Set'})

    return genesets_of_interest_df

def filter_for_genesets_of_interest(all_stat_df, geneset_key='gene_set', only_clustering_databases=True):
    genesets_of_interest_df = get_genesets_of_interest_df(only_clustering_databases=only_clustering_databases)
    genesets_of_interest = list(genesets_of_interest_df['Gene_Set'])
    all_stat_df = all_stat_df[all_stat_df[geneset_key].isin(genesets_of_interest)]
    return all_stat_df


import pandas as pd

def filter_dataframe(df, col_val_dict):
    """
    Filter a DataFrame based on a dictionary where each key is a column name and
    each value can be a single value or a list of acceptable values for that column.

    Parameters:
    - df: The pandas DataFrame to filter.
    - col_val_dict: A dictionary where keys are column names and values can be single values or lists of values to filter by.

    Returns:
    - A filtered DataFrame.
    """
    # Ensure the input is a DataFrame
    if not isinstance(df, pd.DataFrame):
        raise ValueError('The df parameter should be a pandas DataFrame.')

    # Ensure the col_val_dict is a dictionary
    if not isinstance(col_val_dict, dict):
        raise ValueError('The col_val_dict parameter should be a dictionary.')

    # Start with the original DataFrame
    filtered_df = df

    # Apply filters
    for col, val in col_val_dict.items():
        # If the value is not a list, make it a list
        if not isinstance(val, list):
            val = [val]
        filtered_df = filtered_df[filtered_df[col].isin(val)]
    
    return filtered_df

import matplotlib.pyplot as plt
import seaborn as sns

    # def update_legend_style(self, ax, legend_title=None, clean_handles=False, **kwargs):
    #     legend = ax.legend(**kwargs)
    #     #print("AA")
    #     #print(legend.get_title().get_text())
    #     if legend_title == None:
    #         legend.set_title(legend.get_title().get_text(), prop={'size': self.fontsize, 'weight': 'bold'})
    #     else:
    #         legend.set_title(legend_title, prop={'size': self.fontsize , 'weight': 'bold'})
    #     if clean_handles:
    #         handles, labels = ax.get_legend_handles_labels()
    #         new_labels = [label.replace('_', ' ') for label in labels]
    #         ax.legend(handles=handles, labels=new_labels)
    #     return legend

    # def update_legend_style(self, ax, legend_title=None, clean_handles=False, **kwargs):
    # # Check if a legend already exists
    #     legend = ax.get_legend()
    #     if legend is None:
    #         # If no existing legend, create one
    #         legend = ax.legend(**kwargs)

    #     if legend_title is None:
    #         legend.set_title(legend.get_title().get_text(), prop={'size': self.fontsize, 'weight': 'bold'})
    #     else:
    #         legend.set_title(legend_title, prop={'size': self.fontsize , 'weight': 'bold'})
        
    #     if clean_handles:
    #         handles, labels = ax.get_legend_handles_labels()
    #         new_labels = [label.replace('_', ' ') for label in labels]
    #         ax.legend(handles=handles, labels=new_labels)
    #     else:
    #         # Update the existing legend's style without altering the handles and labels
    #         ax.legend(prop={'size': self.fontsize, 'weight': 'bold'})

    #     return legend
    # def set_xticklabels(self, ax, labels, **kwargs):
    #     ax.set_xticklabels(labels, fontsize=self.fontsize, **kwargs)
    # def set_yticklabels(self, ax, labels, **kwargs):
    #     ax.set_xticklabels(labels, fontsize=self.fontsize, **kwargs)
            # You may also want to adjust the spines (the lines noting the boundaries of the plot area)
            # "axes.spines.top": False,   # Disable the top spine
            # "axes.spines.right": False  # Disable the right spine

class FigureStyle:
    def __init__(self, 
                 palette='colorblind', 
                 context='notebook', 
                 style='white', 
                 font='Helvetica', 
                 fontsize=6, 
                 dpi=300,
                 linewidth=0.85,
                 markersize=1,
                 tick_major_pad=0
                ):
        if fontsize < 5 or fontsize > 7:
            fontsize = 6
        self.palette = palette
        self.context = context
        self.style = style
        self.font = font
        self.fontsize = fontsize
        self.dpi = dpi
        self.max_width_inches = 180 / 25.4  # Convert 180 mm to inches
        self.linewidth = linewidth
        self.markersize = markersize
        self.tick_major_pad = tick_major_pad

    def apply(self):
        sns.set_palette(self.palette)
        rc_params = {
            "font.size": self.fontsize,
            "axes.titlesize": self.fontsize + 1,
            "axes.labelsize": self.fontsize,
            "xtick.labelsize": self.fontsize,
            "ytick.labelsize": self.fontsize,
            "legend.fontsize": self.fontsize,  # Set default legend font size
            "lines.linewidth": self.linewidth,  # Default line width for plots
            "axes.linewidth": self.linewidth,   # Default line width for axes borders
            "lines.markersize": self.markersize,
            'xtick.major.pad': self.tick_major_pad,
            'ytick.major.pad': self.tick_major_pad,
            'pdf.fonttype': 42,
            'ps.fonttype': 42,
            "font.family": self.font
        }
        sns.set_context(self.context, rc=rc_params)
        sns.set_style(self.style)
        plt.rcParams.update(rc_params)  # Update plt.rcParams with the specified rc parameters

    def update_legend_style(self, ax, legend_title=None, clean_handles=False, **kwargs):
        legend = ax.legend(**kwargs)
        #print("AA")
        #print(legend.get_title().get_text())
        if legend_title == None:
            legend.set_title(legend.get_title().get_text(), prop={'size': self.fontsize, 'weight': 'bold'})
        else:
            legend.set_title(legend_title, prop={'size': self.fontsize , 'weight': 'bold'})
        if clean_handles:
            handles, labels = ax.get_legend_handles_labels()
            new_labels = [label.replace('_', ' ') for label in labels]
            ax.legend(handles=handles, labels=new_labels)
        return legend

    def set_titles(self, ax, title, fontsize=None):
        if fontsize == None:
            fontsize = self.fontsize+1
        ax.set_title(title, fontsize=fontsize, weight='bold')

    def set_labels(self, ax, xlabel, ylabel):
        ax.set_xlabel(xlabel, fontsize=self.fontsize + 1, weight='bold')
        ax.set_ylabel(ylabel, fontsize=self.fontsize + 1, weight='bold')
        
    def save_figure(self, filename, tight_layout=True, trim_margin=True, **kwargs):
        fig = plt.gcf()
        current_size = fig.get_size_inches()
        if current_size[0] > self.max_width_inches:
            scale_ratio = self.max_width_inches / current_size[0]
            fig.set_size_inches(self.max_width_inches, current_size[1] * scale_ratio)
        if tight_layout:
            plt.tight_layout()

        if trim_margin:
            plt.savefig(filename, dpi=self.dpi, bbox_inches='tight', pad_inches=0.1, **kwargs)
        else:
            # Save without trimming margins. The margins should be set during figure/axes creation.
            plt.savefig(filename, dpi=self.dpi, **kwargs)
    def get_palette_colors(self, n_colors):
        return sns.color_palette(self.palette, n_colors)

def pval2stars(p_value):
        # Define significance levels
    if p_value < 0.00001:
        sig = "***"
    elif p_value < 0.001:
        sig = "**"
    elif p_value < 0.05:
        sig = "*"
    else:
        sig = "ns"  # not significant
    return sig

import json

class JSONDataReader:
    def __init__(self, json_file):
        self.json_file = json_file
        self.data = None
        self.load_json()

    def load_json(self):
        try:
            with open(self.json_file, 'r') as file:
                self.data = json.load(file)
        except FileNotFoundError:
            raise FileNotFoundError(f"The JSON file '{self.json_file}' does not exist.")
        except json.JSONDecodeError:
            raise ValueError(f"The JSON file '{self.json_file}' is not valid JSON.")

    def get(self, attribute_name):

        self.load_json()  # Load JSON if it hasn't been loaded yet
        return self.data.get(attribute_name, None)

    def __getattr__(self, item):
        if self.data is not None:
            return self.data.get(item)
        raise AttributeError(f"'{type(self).__name__}' object has no attribute '{item}'")

class Naming_Json(JSONDataReader):
    def __init__(self):
        config = get_config()
        json_file = f"{config.proj_path}/data/meta/naming/naming.json"
        super().__init__(json_file)  # Call the constructor of the parent class


def get_naming_data():
    naming_dict = get_naming_dict()
    Exp_naming = naming_dict['Exp']
    Comp_naming = naming_dict['Comp']
    Encoding_naming = naming_dict['Encoding']
    Geneset_naming = naming_dict['Geneset']
    def float_to_str(val):
        if pd.isna(val):
            return 'NaN'  # or return str(val) to keep 'nan' as a string
        elif val.is_integer():
            return str(int(val))
        else:
            return str(val)
    njson = Naming_Json()
    Exp_naming['Num_genes'] = Exp_naming['Num_genes'].astype(float).apply(float_to_str)
    Exp_naming['Num_expr_bins'] = Exp_naming['Num_expr_bins'].astype(float).apply(float_to_str)
    Exp_naming['Model'] = Exp_naming['Architecture'].astype(str).replace('nan', '') + '-' + Exp_naming['Objective'].astype(str).replace('nan', 'NA') # + "-" + Exp_naming['Num_genes'].astype(str)
    Exp_naming['Model'] = Exp_naming['Model'].str.replace('-NA', '')
    Exp_naming[njson.get('model_norm_method')] = Exp_naming['Model'] + ' (' + Exp_naming['Normalization'].astype(str).replace('nan', '') + ")"
    Exp_naming[njson.get('model_norm_method')] = Exp_naming[njson.get('model_norm_method')].astype(str).str.replace('()', '', regex=False)
    Exp_naming_dict = Exp_naming.set_index('Exp_name')[njson.get('model_norm_method')].to_dict()
    
    Exp4clf_naming = naming_dict['Exp4clf']
    Clf_for_ROCs = naming_dict['Clf_for_ROCs']
    Exp4clf_dict = Exp4clf_naming.set_index('Exp_name')['Architecture'].to_dict()
    Genesetname_dict = Geneset_naming.set_index('Database_ori_name')['Database'].to_dict()
    Exp_naming['Class'] = Exp_naming['Class'].astype(str).str.replace('"', '', regex=False)
    Exp_naming[njson.get('class_model_norm_method')] = Exp_naming['Class'] + " " + Exp_naming[njson.get('model_norm_method')]
    Exp_naming_class_dict = Exp_naming.set_index('Exp_name')[njson.get('class_model_norm_method')].to_dict()
    return Exp_naming, Comp_naming, Encoding_naming, Geneset_naming, Exp4clf_naming, Clf_for_ROCs, Exp_naming_dict, Exp4clf_dict, Genesetname_dict



import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font
import os
def export_to_excel(dfs_dict, legend, filename, capitalize_headers=True, legend_sheet_name="Table Legend"):
    """
    Export multiple DataFrames to an Excel file, each in a separate sheet.
    
    :param dfs_dict: Dictionary with sheet names as keys and DataFrames as values.
    :param legend: String for the table legend or filename of a text file containing the legend.
    :param filename: Name of the output Excel file.
    :param capitalize_headers: Capitalize the first letter of each column name and replace underscores with spaces.
    :param legend_sheet_name: The name of the sheet that will contain the table legend.
    """
    # Create a new Excel file
    wb = Workbook()
    ws1 = wb.active
    ws1.title = legend_sheet_name

    # Add table legend to the first sheet
    if os.path.isfile(legend):
        with open(legend, 'r') as file:
            for line in file:
                # Splitting the line by tabs and appending as a list to create columns
                ws1.append(line.strip().split('\t'))
    else:
        ws1.append([legend])

    # Add each DataFrame to a separate sheet
    for sheet_name, ori_df in dfs_dict.items():
        df = ori_df.copy()
        if capitalize_headers:
            df.columns = [col.replace('_', ' ').capitalize() for col in df.columns]
        
        ws = wb.create_sheet(title=sheet_name)
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            ws.append(row)
            if r_idx == 1:  # Apply formatting to header row
                for cell in ws[r_idx]:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(wrap_text=True)

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            header_length = len(str(col[0].value))
            adjusted_width = max(max_length, header_length + 2)
            ws.column_dimensions[col[0].column_letter].width = adjusted_width

    # Save the workbook
    wb.save(filename)


